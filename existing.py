#The os module is used to use the clear or cls command from the program itself
import os

#The datetime module is used to keep track of the date and time of the money transactions
from datetime import datetime

#The openpyxl module is being used to deal with the excel sheets that store the personal information and the account details of users
import openpyxl
from openpyxl.utils import get_column_letter 


#this method returns the list of account numbers of existing users
def get_acc_list():
    wb = openpyxl.load_workbook('accounts.xlsx')
    sh = wb.active
    l = []
    for i in range(2, sh.max_row+1):
        l.append(sh[get_column_letter(2)+str(i)].value)
    wb.save('accounts.xlsx')
    return l

#this method returns the security keys of the existing users

def get_key_list():
    wb = openpyxl.load_workbook('accounts.xlsx')
    sh = wb.active
    l = []
    for i in range(2, sh.max_row+1):
        l.append(sh[get_column_letter(6)+str(i)].value)
    wb.save('accounts.xlsx')
    return l


#this method is used to check whether the given account number and security pin are valid and return the row number of the user in the accounts.xlsx file
def check_acc(ac, pin):
    accs = get_acc_list()
    keys = get_key_list()
    if ac in accs:
        i = accs.index(ac)
        if keys[i] == pin:
            return i + 2
        else:
            return 0


#this method prints all the previous transactions  of the given account number
def show_transactions(ac):
    file = 'transacs//' + ac + '.txt'
    f = open(file=file)
    f.seek(0,0)
    os.system('cls')
    print("Transaction History: ")
    print(f.read())
    a = input("Press Enter to continue")
    return

#this method is used to transfer money to an other account and all the date is being saved in the respective account text file in the transacs folder
#also the money details are being updated in the accounts.xlsx file
def transfer(bal, ac):
    name = input("Enter name of recipient: ")
    bank = input("Enter recipient bank name and branch: ")
    acc = input("Enter account number of recipient: ")
    amount = float(input("Enter amount of money to be transferred: "))
    date = datetime.today()
    if amount > bal:
        print("Insufficient balance! The available balance is: Rs.", bal)
        exit(1)
    
    file = 'transacs//' + ac+ '.txt'
    f = open(file, 'a+')
    f.seek(0,0)
    line = str(date) + ' Transfer:\n' + ('\t'*9) +'Sent to: '+name+'\n'+('\t'*9)+' Bank: '+bank +'\n'+('\t'*9)+\
        ' Recipient account number: '+acc+'\n'+('\t'*9)+\
        ' Amount transferred: Rs.'+str(amount) + '\n'+('\t'*9)+' Available Balance: Rs.'+str(bal-amount)+'\n'
    f.write(line)
    f.close()
    print("Transfer successful! Remaining balance: Rs.", (bal-amount))
    return bal - amount


#this method is used to deposit the money into the user account and the details are being stored in the respective text file in transacs folder
#also the money details are being updated in the accounts.xlsx file
def deposit(bal, ac):
    amount = float(input('Enter the amount of money to deposit: Rs.'))
    file = 'transacs//' + ac + '.txt'
    date = datetime.today()
    f = open(file, 'a+')
    f.seek(0,0)
    line = str(date)+' Deposit: Balance before: Rs.'+str(bal)+'\n'+('\t'*9)+'Amount deposited: Rs.'+\
            str(amount)+'\n'+('\t'*9)+'Present balance: Rs.' + str(bal+amount) + '\n'
    f.write(line)
    f.close()
    print("Deposit successful! Remaining balance: Rs.", (bal+amount))
    return bal+amount


#this method is used to withdraw the money from the users account and the details are being stored in the respective text file in transacs folder
#also the money details are being updated in the accounts.xlsx file
def withdraw(bal, ac):
    amount = float(input('Enter the amount of money to withdraw: Rs.'))
    if amount > bal:
        print('Insufficient balance! The available balance is: Rs.', bal)
        exit(1)
    file = 'transacs//' + ac + '.txt'
    date = datetime.today()
    f = open(file, 'a+')
    f.seek(0,0)
    line = str(date)+' Withdrawl: Balance before: Rs.'+str(bal)+str(bal)+'\n'+('\t'*9)+'    Amount withdrawn: Rs.'+\
        str(amount) + '\n'+('\t'*9)+'    Present balance: Rs.' + str(bal-amount) + '\n'
    f.write(line)
    f.close()
    print("Withdrawl successful! Remaining balance: Rs.", (bal-amount))
    return bal-amount
    

#this method is the heart of the program. This method runs all the above mentioned definitions

def user():
    os.system('cls')

    #first the account number and security pin are taken as input
    ac = input('Enter account number: ')
    pin = input('Enter security pin: ')

    #the authenticity of the account number and pin are being checked and the row number of the acc number in the accounts.xlsx file is being returned
    row = check_acc(ac, pin)

    #if the details are found to be wrong we prompt the user to try again with a different account number and security pin or to exit
    while row == 0:
        os.system('cls')
        print('Invalid account number or security pin!Enter \'e\' to exit or try again')
        ac = input('Enter account number: ')
        if ac == 'e':
            exit(1)
        pin = input('Enter security pin: ')
        row = check_acc(ac, pin)
    
    os.system('cls')


    #once the authenticity of the account number and pin are verified the following prompts will be displayed for the user
    print('LOGIN SUCCESSFUL!\n\n')

    #the accounts.xlsx file is being opened for the updating of the balance amount of the user if required 
    wb = openpyxl.load_workbook(filename = 'accounts.xlsx')
    sh = wb['Sheet1']

    #this is an infinite loop run until the user chooses to exit
    while True:
        print('1.Check balance\n2.Deposit money\n3.Withdraw money\n4.Transfer money\n5.Show Transaction History\n6.Quit')
        choice = int(input('Enter your choice: '))

        if  choice == 1:
            #printing the available balance in the users account
            print('Available balance: Rs.', sh.cell(row, 7).value)

        elif choice == 2:
            #the deposit() method is being run and the returned value is the final balance which is being updated in the excel file
            result = deposit(sh.cell(row, 7).value, ac)
            sh.cell(row, 7, value = result)

        elif choice == 3:
            #the withdraw() method is being run and the returned value is the final balance which is being updated in the excel file
            result = withdraw(sh.cell(row, 7).value, ac)
            sh.cell(row, 7, value=result)

        elif choice == 4:
            #the transfer() method is being run and the returned value is the final balance which is being updated in the excel file           
            result = transfer(sh.cell(row, 7).value, ac)
            sh.cell(row, 7, value=result)

        elif choice == 5:
            #the show_transactions() method is being used to print all previous transactions of the user
            show_transactions(ac)

        elif choice == 6:
            #the excel file is being saved 
            #the program is being exited
            wb.save('accounts.xlsx')
            exit(0)
        os.system('cls')
